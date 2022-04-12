import openpyxl

book = openpyxl.load_workbook("C:\\Users\\umayadav\\PycharmProjects\\TestDataPython.xlsx")
sheet = book.active
dict = {} #declare dictionary
print(sheet.max_row)  # total number of rows
print(sheet.max_column)  # total number of columns
for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "Test2":
        for j in range(2, sheet.max_column + 1):
            # print((sheet.cell(row=i, column=j)).value)
            dict[(sheet.cell(row=1, column=j)).value] = (sheet.cell(row=i, column=j)).value
print(dict)
